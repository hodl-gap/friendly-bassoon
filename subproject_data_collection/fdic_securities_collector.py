"""
FDIC Bank Securities Data Collector

Downloads HTM/AFS securities data (holdings + unrealized gains/losses) from two sources:
1. FDIC QBP Excel spreadsheet (aggregate, has AFS amortized cost + unrealized)
2. FDIC BankFind API (per-bank, summed; has HTM cost, AFS FV, total, total MV)

Merges both into a single wide CSV and exports individual date,value series.

Usage:
    python subproject_data_collection/fdic_securities_collector.py
    python subproject_data_collection/fdic_securities_collector.py --api-only
    python subproject_data_collection/fdic_securities_collector.py --merge-only
    python subproject_data_collection/fdic_securities_collector.py --export-only
"""

import argparse
import csv
import json
import random
import time
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

DATA_DIR = Path(__file__).parent / "data"
RAW_DIR = DATA_DIR / "fdic_raw"
OUTPUT_CSV = DATA_DIR / "fdic_securities.csv"
CSV_SERIES_DIR = DATA_DIR / "csv_series"
PROGRESS_FILE = RAW_DIR / "api_progress.json"
API_CSV = RAW_DIR / "fdic_api_aggregates.csv"
QBP_XLSX = RAW_DIR / "qbp_balance_sheet.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36",
}

# FDIC BankFind API
API_BASE = "https://api.fdic.gov/banks/financials"
API_FIELDS = "REPDTE,CERT,SCHA,SCAF,SC,SCMV"
API_LIMIT = 5000

# Quarter labels from 1994Q1 to present
QUARTER_NAMES = {1: "first", 2: "second", 3: "third", 4: "fourth"}

# QBP Excel row indices (1-indexed) for Balance Sheet securities data
QBP_ROWS = {
    38: "afs_fair_value",
    39: "afs_amortized_cost",
    40: "afs_unrealized",
    41: "htm_fair_value",
    42: "htm_amortized_cost",
    43: "htm_unrealized",
    44: "total_unrealized",
}

# Wide CSV columns
WIDE_FIELDNAMES = [
    "date",
    "afs_fair_value", "afs_amortized_cost", "afs_unrealized",
    "htm_fair_value", "htm_amortized_cost", "htm_unrealized",
    "total_unrealized",
    "api_scha", "api_scaf", "api_sc", "api_scmv",
    "api_htm_unrealized", "api_bank_count",
    "source",
]

# Individual series to export (column name → output filename)
EXPORT_SERIES = {
    "afs_fair_value": "fdic_afs_fair_value",
    "afs_amortized_cost": "fdic_afs_amortized_cost",
    "afs_unrealized": "fdic_afs_unrealized",
    "htm_fair_value": "fdic_htm_fair_value",
    "htm_amortized_cost": "fdic_htm_amortized_cost",
    "htm_unrealized": "fdic_htm_unrealized",
    "total_unrealized": "fdic_total_unrealized",
    "api_sc": "fdic_total_securities",
    "api_scmv": "fdic_total_securities_mv",
}


# ── Helpers ────────────────────────────────────────────────────────────────


def _quarter_end_date(year: int, quarter: int) -> str:
    """Return YYYY-MM-DD for the last day of a quarter."""
    month = quarter * 3
    if month == 3:
        return f"{year}-03-31"
    elif month == 6:
        return f"{year}-06-30"
    elif month == 9:
        return f"{year}-09-30"
    else:
        return f"{year}-12-31"


def _repdte(year: int, quarter: int) -> str:
    """Return FDIC REPDTE format YYYYMMDD for quarter end."""
    month = quarter * 3
    day = 31 if month in (3, 12) else 30
    return f"{year}{month:02d}{day:02d}"


def _generate_quarters(start_year=1994, start_q=1):
    """Generate (year, quarter) tuples from start to present."""
    now = datetime.now()
    current_year = now.year
    current_quarter = (now.month - 1) // 3 + 1
    quarters = []
    for year in range(start_year, current_year + 1):
        for q in range(1, 5):
            if year == start_year and q < start_q:
                continue
            if year == current_year and q > current_quarter:
                break
            quarters.append((year, q))
    return quarters


# ── Step 1: Download QBP Excel ─────────────────────────────────────────────


def _find_qbp_url() -> str | None:
    """Find the latest available QBP Excel URL by trying HEAD requests backwards."""
    now = datetime.now()
    year = now.year
    quarter = (now.month - 1) // 3 + 1

    # Try from current quarter backwards for up to 8 quarters
    attempts = 0
    y, q = year, quarter
    while attempts < 8:
        name = QUARTER_NAMES[q]
        url = f"https://www.fdic.gov/quarterly-banking-profile/qbp-time-series-spreadsheet-{name}-quarter-{y}.xlsx"
        try:
            req = Request(url, method="HEAD", headers=HEADERS)
            with urlopen(req, timeout=15) as resp:
                if resp.status == 200:
                    print(f"  [found] QBP Excel: {name}-quarter-{y}")
                    return url
        except (HTTPError, URLError):
            pass
        # Previous quarter
        q -= 1
        if q == 0:
            q = 4
            y -= 1
        attempts += 1

    return None


def download_qbp() -> dict[str, dict]:
    """Download QBP Excel and parse Balance Sheet securities rows.

    Returns dict keyed by date string (YYYY-MM-DD), values are dicts with
    afs_fair_value, afs_amortized_cost, afs_unrealized, htm_fair_value,
    htm_amortized_cost, htm_unrealized, total_unrealized.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    # Download if not cached
    if not QBP_XLSX.exists():
        url = _find_qbp_url()
        if not url:
            print("  [error] Could not find QBP Excel URL")
            return {}
        print(f"  [download] {url}")
        req = Request(url, headers=HEADERS)
        try:
            with urlopen(req, timeout=60) as resp:
                data = resp.read()
            QBP_XLSX.write_bytes(data)
            print(f"  [ok] {QBP_XLSX.name} ({len(data):,} bytes)")
        except (HTTPError, URLError) as e:
            print(f"  [error] Download failed: {e}")
            return {}
    else:
        print(f"  [skip] {QBP_XLSX.name} already exists ({QBP_XLSX.stat().st_size:,} bytes)")

    # Parse with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(QBP_XLSX, read_only=True, data_only=True)
    ws = wb["Balance Sheet"]

    # Row 6 = quarter labels. Read all cells to find data range.
    row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]

    # Find column indices with quarter labels (format like "1994Q1" or similar)
    quarter_cols = {}  # col_index → date string
    for col_idx, val in enumerate(row6):
        if val is None:
            continue
        val_str = str(val).strip()
        # Try parsing various quarter label formats
        date = _parse_quarter_label(val_str)
        if date:
            quarter_cols[col_idx] = date

    if not quarter_cols:
        print("  [error] Could not find quarter labels in QBP row 6")
        wb.close()
        return {}

    print(f"  [parse] Found {len(quarter_cols)} quarters in QBP "
          f"({min(quarter_cols.values())} to {max(quarter_cols.values())})")

    # Read securities rows
    result = {}
    for row_num, field_name in QBP_ROWS.items():
        row_data = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
        for col_idx, date in quarter_cols.items():
            if date not in result:
                result[date] = {}
            val = row_data[col_idx] if col_idx < len(row_data) else None
            if val is not None:
                try:
                    result[date][field_name] = float(val)
                except (ValueError, TypeError):
                    pass

    wb.close()

    # Filter out dates with no data
    result = {d: v for d, v in result.items() if v}
    print(f"  [ok] Parsed {len(result)} quarters from QBP")
    return result


def _parse_quarter_label(label: str) -> str | None:
    """Parse quarter label to YYYY-MM-DD date string.

    Handles formats like: "1994Q1", "1994 Q1", "Q1 1994", "1Q94",
    "Mar-94", "2024Q4", etc.
    """
    label = label.strip()

    # Format: YYYYQn or YYYY Qn
    for sep in ("Q", " Q", "q", " q"):
        if sep in label:
            parts = label.split(sep.strip())
            if len(parts) == 2:
                try:
                    left, right = parts[0].strip(), parts[1].strip()
                    # Could be "1994Q1" or "Q1 1994"
                    if len(left) == 4 and left.isdigit():
                        year = int(left)
                        q = int(right[0])  # First char of right part
                    elif len(right) >= 4 and right[:4].isdigit():
                        year = int(right[:4])
                        q = int(left[-1]) if left else int(right[0])
                    else:
                        continue
                    if 1 <= q <= 4 and 1990 <= year <= 2030:
                        return _quarter_end_date(year, q)
                except (ValueError, IndexError):
                    continue

    return None


# ── Step 2: Fetch API data ─────────────────────────────────────────────────


def _load_api_progress() -> dict:
    """Load API fetch progress."""
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text())
    return {"fetched_quarters": [], "aggregates": {}}


def _save_api_progress(progress: dict):
    PROGRESS_FILE.write_text(json.dumps(progress, indent=2))


def _fetch_quarter_api(year: int, quarter: int) -> dict | None:
    """Fetch all banks for one quarter from FDIC API, return aggregates.

    Returns dict with scha, scaf, sc, scmv (in $K), bank_count.
    Handles pagination for quarters with >5000 banks.
    """
    repdte = _repdte(year, quarter)
    offset = 0
    totals = {"scha": 0.0, "scaf": 0.0, "sc": 0.0, "scmv": 0.0, "bank_count": 0}

    while True:
        url = (f"{API_BASE}?filters=REPDTE:{repdte}"
               f"&fields={API_FIELDS}&limit={API_LIMIT}&offset={offset}")
        req = Request(url, headers=HEADERS)
        try:
            with urlopen(req, timeout=30) as resp:
                data = json.loads(resp.read())
        except (HTTPError, URLError, json.JSONDecodeError) as e:
            print(f"  [error] API {year}Q{quarter}: {e}")
            return None

        rows = data.get("data", [])
        if not rows:
            break

        for row in rows:
            d = row.get("data", {})
            for field, key in [("SCHA", "scha"), ("SCAF", "scaf"),
                               ("SC", "sc"), ("SCMV", "scmv")]:
                val = d.get(field)
                if val is not None:
                    try:
                        totals[key] += float(val)
                    except (ValueError, TypeError):
                        pass
            totals["bank_count"] += 1

        # Check if we need another page
        total_count = data.get("totals", {}).get("count", 0)
        offset += len(rows)
        if offset >= total_count or len(rows) < API_LIMIT:
            break

        time.sleep(random.uniform(0.3, 0.8))

    if totals["bank_count"] == 0:
        return None

    return totals


def fetch_api_data() -> dict[str, dict]:
    """Fetch all quarters from FDIC API. Resumable via progress file.

    Returns dict keyed by date string, values are dicts with
    api_scha, api_scaf, api_sc, api_scmv (in $M), api_htm_unrealized, api_bank_count.
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    progress = _load_api_progress()
    fetched_set = set(progress["fetched_quarters"])
    quarters = _generate_quarters()
    new_count = 0

    print(f"  [resume] {len(fetched_set)} quarters already fetched")

    for year, q in quarters:
        key = f"{year}Q{q}"
        if key in fetched_set:
            continue

        totals = _fetch_quarter_api(year, q)
        if totals is None:
            # No data for this quarter — mark as fetched but empty
            fetched_set.add(key)
            progress["fetched_quarters"] = sorted(fetched_set)
            continue

        date = _quarter_end_date(year, q)

        # Convert $K to $M
        scha_m = totals["scha"] / 1000
        scaf_m = totals["scaf"] / 1000
        sc_m = totals["sc"] / 1000
        scmv_m = totals["scmv"] / 1000

        # HTM unrealized = (total MV - AFS FV) - HTM cost
        htm_unreal = (scmv_m - scaf_m) - scha_m if scmv_m and scaf_m else None

        progress["aggregates"][date] = {
            "api_scha": round(scha_m, 1),
            "api_scaf": round(scaf_m, 1),
            "api_sc": round(sc_m, 1),
            "api_scmv": round(scmv_m, 1),
            "api_htm_unrealized": round(htm_unreal, 1) if htm_unreal is not None else None,
            "api_bank_count": totals["bank_count"],
        }

        fetched_set.add(key)
        progress["fetched_quarters"] = sorted(fetched_set)
        new_count += 1

        if new_count % 10 == 0:
            print(f"  [progress] {new_count} new quarters fetched (at {key}, "
                  f"{totals['bank_count']} banks)")
            _save_api_progress(progress)

        time.sleep(random.uniform(0.3, 0.8))

    _save_api_progress(progress)

    # Also save intermediate CSV
    if progress["aggregates"]:
        api_fields = ["date", "api_scha", "api_scaf", "api_sc", "api_scmv",
                       "api_htm_unrealized", "api_bank_count"]
        sorted_dates = sorted(progress["aggregates"].keys())
        with open(API_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=api_fields)
            writer.writeheader()
            for date in sorted_dates:
                row = {"date": date, **progress["aggregates"][date]}
                # Replace None with empty
                row = {k: ("" if v is None else v) for k, v in row.items()}
                writer.writerow(row)
        print(f"  [saved] {API_CSV} ({len(sorted_dates)} rows)")

    print(f"  [done] {new_count} new quarters fetched, "
          f"{len(progress['aggregates'])} total quarters with data")

    return progress["aggregates"]


# ── Step 3: Merge ──────────────────────────────────────────────────────────


def merge(qbp_data: dict | None = None, api_data: dict | None = None):
    """Merge QBP + API data into single wide CSV.

    If data dicts not provided, loads from cached files.
    """
    # Load QBP data from xlsx if not provided
    if qbp_data is None:
        if QBP_XLSX.exists():
            print("  [load] Re-parsing QBP Excel...")
            qbp_data = download_qbp()
        else:
            qbp_data = {}

    # Load API data from progress file if not provided
    if api_data is None:
        progress = _load_api_progress()
        api_data = progress.get("aggregates", {})
        if api_data:
            print(f"  [load] Loaded {len(api_data)} API quarters from progress file")

    # Collect all dates
    all_dates = sorted(set(list(qbp_data.keys()) + list(api_data.keys())))

    if not all_dates:
        print("  [error] No data to merge")
        return

    # Merge
    rows = []
    mismatches = []
    for date in all_dates:
        qbp = qbp_data.get(date, {})
        api = api_data.get(date, {})

        has_qbp = bool(qbp)
        has_api = bool(api)

        if has_qbp and has_api:
            source = "both"
        elif has_qbp:
            source = "qbp"
        else:
            source = "api"

        row = {"date": date, "source": source}

        # QBP fields
        for field in ["afs_fair_value", "afs_amortized_cost", "afs_unrealized",
                       "htm_fair_value", "htm_amortized_cost", "htm_unrealized",
                       "total_unrealized"]:
            val = qbp.get(field)
            row[field] = round(val, 1) if val is not None else ""

        # API fields
        for field in ["api_scha", "api_scaf", "api_sc", "api_scmv",
                       "api_htm_unrealized", "api_bank_count"]:
            val = api.get(field)
            row[field] = val if val is not None else ""

        # Cross-verify HTM cost between sources
        if has_qbp and has_api:
            qbp_htm_cost = qbp.get("htm_amortized_cost")
            api_htm_cost = api.get("api_scha")
            if qbp_htm_cost and api_htm_cost and qbp_htm_cost != 0:
                pct_diff = abs(qbp_htm_cost - api_htm_cost) / abs(qbp_htm_cost) * 100
                if pct_diff > 5:
                    mismatches.append((date, qbp_htm_cost, api_htm_cost, pct_diff))

        rows.append(row)

    # Write wide CSV
    with open(OUTPUT_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=WIDE_FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in WIDE_FIELDNAMES})

    # Summary
    both_count = sum(1 for r in rows if r["source"] == "both")
    qbp_only = sum(1 for r in rows if r["source"] == "qbp")
    api_only = sum(1 for r in rows if r["source"] == "api")
    print(f"  [merged] {OUTPUT_CSV}")
    print(f"  [stats] {len(rows)} quarters ({both_count} both, "
          f"{qbp_only} qbp-only, {api_only} api-only)")
    print(f"  [stats] Date range: {all_dates[0]} to {all_dates[-1]}")

    if mismatches:
        print(f"  [warn] {len(mismatches)} HTM cost mismatches >5% between QBP and API:")
        for date, qbp_val, api_val, pct in mismatches[:5]:
            print(f"    {date}: QBP={qbp_val:,.0f} vs API={api_val:,.0f} ({pct:.1f}%)")
        if len(mismatches) > 5:
            print(f"    ... and {len(mismatches) - 5} more")

    # Spot-check key values
    for row in rows:
        if row["date"] == "2022-09-30" and row.get("total_unrealized"):
            print(f"  [check] Q3 2022 total_unrealized = {row['total_unrealized']:,} "
                  f"(expected ~-688,174)")
        if row["date"] == all_dates[-1] and row.get("total_unrealized"):
            print(f"  [check] Latest ({row['date']}) total_unrealized = "
                  f"{row['total_unrealized']:,}")


# ── Step 4: Export long-form CSVs ──────────────────────────────────────────


def export_long_form():
    """Export wide CSV into individual date,value series files."""
    if not OUTPUT_CSV.exists():
        print(f"  [error] Wide CSV not found: {OUTPUT_CSV}")
        print("  [hint] Run without --export-only first to fetch and merge data")
        return

    CSV_SERIES_DIR.mkdir(parents=True, exist_ok=True)

    rows = []
    with open(OUTPUT_CSV, newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    exported = 0
    for col_name, file_name in EXPORT_SERIES.items():
        pairs = []
        for row in rows:
            val = row.get(col_name, "").strip()
            if val:
                pairs.append((row["date"], val))

        if not pairs:
            continue

        out_path = CSV_SERIES_DIR / f"{file_name}.csv"
        with open(out_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["date", "value"])
            for date, value in pairs:
                writer.writerow([date, value])

        exported += 1
        print(f"  [export] {file_name}.csv ({len(pairs)} rows)")

    print(f"  [done] Exported {exported} series to {CSV_SERIES_DIR}")


# ── CLI ────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="FDIC Bank Securities Data Collector")
    parser.add_argument("--api-only", action="store_true",
                        help="Skip QBP download, only fetch API data")
    parser.add_argument("--merge-only", action="store_true",
                        help="Re-merge existing data files")
    parser.add_argument("--export-only", action="store_true",
                        help="Re-export long-form CSVs from existing wide CSV")
    args = parser.parse_args()

    RAW_DIR.mkdir(parents=True, exist_ok=True)

    if args.export_only:
        print("[export] Exporting long-form CSVs from wide CSV...")
        export_long_form()
        return

    if args.merge_only:
        print("[merge] Re-merging existing data...")
        merge()
        print("[step 4] Exporting long-form CSVs...")
        export_long_form()
        return

    qbp_data = {}
    if not args.api_only:
        print("[step 1] Downloading QBP Excel...")
        qbp_data = download_qbp()

    print("[step 2] Fetching FDIC API data...")
    api_data = fetch_api_data()

    print("[step 3] Merging all data...")
    merge(qbp_data, api_data)

    print("[step 4] Exporting long-form CSVs...")
    export_long_form()

    print(f"[done] Output: {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
